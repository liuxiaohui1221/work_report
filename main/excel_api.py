#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 导入/导出模块 - 任务文件管理
支持上传Excel、编辑、保存，同时记录变更历史
"""

import os
import json
from datetime import datetime
from flask import Blueprint, request, jsonify
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from werkzeug.utils import secure_filename

from main.models import (
    TaskDAO, TaskHistoryDAO, DeveloperDAO, ProjectDAO, TaskService
)


# 创建蓝图
excel_bp = Blueprint('excel', __name__, url_prefix='/api/excel')

# 上传目录
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "input", "tasks")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

ALLOWED_EXTENSIONS = {'xlsx'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ========== Excel 解析 ==========

def parse_task_excel(filepath: str) -> dict:
    """
    解析任务Excel文件
    返回格式：
    {
        'headers': [...],
        'rows': [...],
        'tasks': [...]  # 解析后的任务列表
    }
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb.active

    # 获取所有行
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {'headers': [], 'rows': [], 'tasks': []}

    headers = [str(c).strip() if c else "" for c in rows[0]]

    # 建立列映射
    col_map = {}
    for idx, h in enumerate(headers):
        h_lower = h.lower()
        if any(keyword in h for keyword in ['模块', '名称', '项目']):
            col_map['module'] = idx
        if any(keyword in h for keyword in ['功能', '描述', '任务']):
            col_map['description'] = idx
        if any(keyword in h for keyword in ['状态']):
            col_map['status'] = idx
        if any(keyword in h for keyword in ['负责人', '指派']):
            col_map['assignee'] = idx
        if any(keyword in h for keyword in ['截止', '期限', '日期', 'due']):
            col_map['due_date'] = idx
        if any(keyword in h for keyword in ['优先级', 'priority']):
            col_map['priority'] = idx
        if any(keyword in h for keyword in ['依赖', '前置']):
            col_map['depends_on'] = idx
        if any(keyword in h for keyword in ['ID', 'id']):
            col_map['task_id'] = idx

    # 解析任务数据
    tasks = []
    for row_idx, row in enumerate(rows[1:], start=2):
        task = {'row': row_idx}

        for col_name, col_idx in col_map.items():
            if col_idx < len(row):
                value = row[col_idx]
                if value is not None:
                    task[col_name] = str(value).strip()
                else:
                    task[col_name] = ""
            else:
                task[col_name] = ""

        # 只添加有内容的行
        if task.get('description') or task.get('module'):
            tasks.append(task)

    return {
        'headers': headers,
        'rows': rows[1:],
        'tasks': tasks,
        'row_count': len(rows) - 1,
        'sheet_name': sheet.title
    }


def create_task_excel(tasks: list, filepath: str = None) -> str:
    """
    创建任务Excel文件
    返回文件路径
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "任务列表"

    # 设置样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 写入表头
    headers = ['ID', '模块', '功能描述', '状态', '优先级', '负责人', '截止日期', '依赖前置']
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据
    for row_idx, task in enumerate(tasks, start=2):
        # ID
        sheet.cell(row=row_idx, column=1, value=task.get('id', ''))
        # 模块
        sheet.cell(row=row_idx, column=2, value=task.get('module', ''))
        # 功能描述
        sheet.cell(row=row_idx, column=3, value=task.get('title', task.get('description', '')))
        # 状态
        sheet.cell(row=row_idx, column=4, value=task.get('status', ''))
        # 优先级
        sheet.cell(row=row_idx, column=5, value=task.get('priority', ''))
        # 负责人
        sheet.cell(row=row_idx, column=6, value=task.get('assignee_name', task.get('assignee', '')))
        # 截止日期
        sheet.cell(row=row_idx, column=7, value=task.get('due_date', ''))
        # 依赖
        sheet.cell(row=row_idx, column=8, value=task.get('depends_on_title', task.get('depends_on', '')))

    # 设置列宽
    widths = [8, 15, 40, 12, 10, 15, 15, 20]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + idx)].width = width

    # 保存
    if not filepath:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = os.path.join(UPLOAD_FOLDER, f"tasks_export_{timestamp}.xlsx")

    wb.save(filepath)
    return filepath


# ========== API 端点 ==========

@excel_bp.route('/upload', methods=['POST'])
def upload_excel():
    """上传Excel文件"""
    if 'file' not in request.files:
        return jsonify({'error': '没有文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '没有选择文件'}), 400

    if not allowed_file(file.filename):
        return jsonify({'error': '只支持.xlsx文件'}), 400

    # 保存文件
    filename = secure_filename(file.filename)
    project_name = request.form.get('project', 'default')
    project_dir = os.path.join(UPLOAD_FOLDER, project_name)
    os.makedirs(project_dir, exist_ok=True)

    filepath = os.path.join(project_dir, filename)
    file.save(filepath)

    # 解析Excel
    try:
        data = parse_task_excel(filepath)
        return jsonify({
            'filename': filename,
            'project': project_name,
            'filepath': filepath,
            'row_count': data['row_count'],
            'tasks_preview': data['tasks'][:20],  # 预览前20条
            'headers': data['headers']
        })
    except Exception as e:
        return jsonify({'error': f'解析Excel失败: {str(e)}'}), 400


@excel_bp.route('/parse/<filename>', methods=['GET'])
def parse_excel(filename):
    """解析已上传的Excel文件"""
    # 解码URL编码的文件名
    filename = secure_filename(filename)

    # 搜索文件
    filepath = None
    for root, dirs, files in os.walk(UPLOAD_FOLDER):
        if filename in files:
            filepath = os.path.join(root, filename)
            break

    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'}), 404

    try:
        data = parse_task_excel(filepath)
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': f'解析失败: {str(e)}'}), 400


@excel_bp.route('/list', methods=['GET'])
def list_files():
    """列出已上传的Excel文件"""
    files = []
    if os.path.exists(UPLOAD_FOLDER):
        for root, dirs, filenames in os.walk(UPLOAD_FOLDER):
            for filename in filenames:
                if filename.endswith('.xlsx'):
                    full_path = os.path.join(root, filename)
                    rel_path = os.path.relpath(full_path, UPLOAD_FOLDER)

                    # 获取文件修改时间
                    mtime = os.path.getmtime(full_path)
                    modified = datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M')

                    # 获取项目名（第一级目录）
                    project = rel_path.split(os.sep)[0] if os.sep in rel_path else 'default'

                    files.append({
                        'name': filename,
                        'path': full_path,
                        'relative_path': rel_path,
                        'project': project,
                        'size': os.path.getsize(full_path),
                        'modified': modified
                    })

    # 按修改时间排序，最新的在前
    files.sort(key=lambda x: x['modified'], reverse=True)
    return jsonify(files)


@excel_bp.route('/import', methods=['POST'])
def import_tasks():
    """
    从Excel导入任务到数据库
    请求体：
    {
        'project_id': 1,
        'filepath': '/path/to/file.xlsx',
        'developer_id': 1,  # 导人
        'mapping': {  // 列映射
            'module': 0,
            'title': 1,
            'status': 2,
            'assignee': 3,
            'due_date': 4
        }
    }
    """
    data = request.json
    project_id = data.get('project_id')
    filepath = data.get('filepath')
    developer_id = data.get('developer_id')  # 执行导入的人
    mapping = data.get('mapping', {})

    if not project_id:
        return jsonify({'error': '缺少项目ID'}), 400

    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'}), 404

    try:
        # 解析Excel
        excel_data = parse_task_excel(filepath)

        created_tasks = []
        errors = []

        for row_idx, row_data in enumerate(excel_data['tasks']):
            try:
                # 根据mapping提取字段
                title = row_data.get(mapping.get('title', 'description') or 'description', '')
                module = row_data.get(mapping.get('module', 'module') or 'module', '')
                status = row_data.get(mapping.get('status', 'status') or 'status', '待认领')
                priority = row_data.get(mapping.get('priority', 'priority') or 'priority', '中')
                assignee_name = row_data.get(mapping.get('assignee', 'assignee') or 'assignee', '')
                due_date = row_data.get(mapping.get('due_date', 'due_date') or 'due_date', '')

                # 如果没有标题，跳过
                if not title:
                    continue

                # 查找负责人
                assignee_id = None
                if assignee_name:
                    developers = DeveloperDAO.get_by_project(project_id)
                    for dev in developers:
                        if assignee_name in dev['name'] or dev['name'] in assignee_name:
                            assignee_id = dev['id']
                            break

                # 创建任务
                task = TaskService.create_task(
                    project_id=project_id,
                    title=title,
                    description=row_data.get('description', ''),
                    status=status,
                    priority=priority,
                    assignee_id=assignee_id,
                    reporter_id=developer_id,
                    due_date=due_date if due_date else None,
                    module=module
                )
                created_tasks.append(task)

            except Exception as e:
                errors.append({
                    'row': row_idx + 2,  # Excel行号（从2开始，因为1是表头）
                    'data': row_data,
                    'error': str(e)
                })

        # 记录导入历史
        if created_tasks:
            TaskHistoryDAO.add(
                task_id=created_tasks[0]['id'] if created_tasks else 0,
                field_name='import',
                old_value='',
                new_value=f'从Excel导入 {len(created_tasks)} 个任务',
                changed_by=developer_id,
                reason=f'导入文件: {os.path.basename(filepath)}'
            )

        return jsonify({
            'success': True,
            'created_count': len(created_tasks),
            'error_count': len(errors),
            'tasks': created_tasks,
            'errors': errors
        })

    except Exception as e:
        return jsonify({'error': f'导入失败: {str(e)}'}), 400


@excel_bp.route('/export', methods=['POST'])
def export_tasks():
    """
    导出任务为Excel
    请求体：
    {
        'project_id': 1,
        'tasks': [...]  // 可选，如果不提供则导出项目所有任务
    }
    """
    data = request.json
    project_id = data.get('project_id')
    tasks_data = data.get('tasks')

    if not project_id:
        return jsonify({'error': '缺少项目ID'}), 400

    try:
        # 获取任务数据
        if tasks_data:
            tasks = tasks_data
        else:
            tasks = TaskDAO.get_by_project(project_id)

        # 创建Excel
        filepath = create_task_excel(tasks)

        return jsonify({
            'success': True,
            'filepath': filepath,
            'filename': os.path.basename(filepath),
            'task_count': len(tasks)
        })

    except Exception as e:
        return jsonify({'error': f'导出失败: {str(e)}'}), 400


@excel_bp.route('/save', methods=['POST'])
def save_excel_edits():
    """
    保存Excel编辑（修改任务）
    请求体：
    {
        'project_id': 1,
        'changes': [
            {
                'task_id': 1,
                'field': 'status',
                'old_value': '待认领',
                'new_value': '开发中'
            },
            ...
        ],
        'developer_id': 1
    }
    """
    data = request.json
    project_id = data.get('project_id')
    changes = data.get('changes', [])
    developer_id = data.get('developer_id')

    if not changes:
        return jsonify({'error': '没有修改'}), 400

    updated_tasks = []
    errors = []

    for change in changes:
        task_id = change.get('task_id')
        field = change.get('field')
        old_value = change.get('old_value')
        new_value = change.get('new_value')

        if not all([task_id, field, new_value is not None]):
            errors.append({'change': change, 'error': '缺少必要字段'})
            continue

        try:
            # 更新任务
            task = TaskService.update_task(
                task_id,
                changed_by=developer_id,
                **{field: new_value}
            )
            updated_tasks.append({
                'task_id': task_id,
                'field': field,
                'old_value': old_value,
                'new_value': new_value,
                'task': task
            })
        except Exception as e:
            errors.append({
                'task_id': task_id,
                'field': field,
                'error': str(e)
            })

    return jsonify({
        'success': True,
        'updated_count': len(updated_tasks),
        'error_count': len(errors),
        'updated_tasks': updated_tasks,
        'errors': errors
    })


@excel_bp.route('/history/<int:project_id>', methods=['GET'])
def get_project_task_history(project_id):
    """获取项目的任务变更历史"""
    # 获取项目的所有任务
    tasks = TaskDAO.get_by_project(project_id)

    all_history = []
    for task in tasks:
        history = TaskHistoryDAO.get_by_task(task['id'])
        for h in history:
            h['task_title'] = task.get('title', '')
        all_history.extend(history)

    # 按时间排序，最新的在前
    all_history.sort(key=lambda x: x.get('created_at', ''), reverse=True)

    return jsonify({
        'total': len(all_history),
        'history': all_history[:100]  # 限制返回100条
    })


@excel_bp.route('/template', methods=['GET'])
def download_template():
    """下载任务导入模板"""
    # 创建空模板
    tasks = []
    filepath = create_task_excel(tasks, os.path.join(UPLOAD_FOLDER, "task_template.xlsx"))

    return jsonify({
        'filepath': filepath,
        'filename': 'task_template.xlsx'
    })