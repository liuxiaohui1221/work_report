#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
自动化日报生成脚本（团队版）
功能：整合本地规划文件 + 团队成员当天Git提交记录和代码变更，按人按项目生成日报。
使用前：
1. 安装依赖：pip install openai requests openpyxl python-dotenv anthropic
2. 修改 config.py 中的团队成员配置
3. 修改 prompts 目录下的 md 文件可自定义 prompt
4. 运行：python daily_report.py
"""

import os
import subprocess
from datetime import datetime, timedelta
from dotenv import load_dotenv
import anthropic
import openpyxl
from config import (
    PROJECT1_NAME, PROJECT1_REPO_PATH,
    PROJECT2_NAME, PROJECT2_REPO_PATH,
    PLAN_DATA_DIR, PLATFORM_DIR, AGENT_DIR,
    MAX_DIFF_LINES, MODEL_NAME, MAX_TOKENS, TEMPERATURE,
    TEAM_MEMBERS, TEAM_MEMBER_SOURCE, DEFAULT_ROLE,
    DAILY_REPORT_MODE, DAILY_REPORT_DAYS, MORNING_HOUR, EVENING_HOUR,
    EMAIL_ENABLED, EMAIL_SMTP_USER, TEAM_MEMBER_SEND_EMAIL, TEAM_MEMBER_EMAIL_MODE,
    PROJECT_REPORT_RECIPIENTS, DEFAULT_PROJECT_RECIPIENTS,
    PENDING_STATUSES, COMPLETED_STATUS, MAX_PENDING_TASKS
)
from email_sender import send_personal_report_email, send_project_report_email, confirm_email_sender_and_recipients, set_email_sender, select_recipients_from_git_authors, reset_email_rate_limit, select_project_report_recipients

# Prompt 文件路径
PROMPTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "prompts")
PERSONAL_DAILY_PROMPT_FILE = os.path.join(PROMPTS_DIR, "personal_daily_prompt.md")
PROJECT_DAILY_PROMPT_FILE = os.path.join(PROMPTS_DIR, "project_daily_prompt.md")

load_dotenv()

MINIMAX_API_KEY = os.getenv("MINIMAX_API_KEY")
MINIMAX_BASE_URL = os.getenv("MINIMAX_BASE_URL", "https://api.minimaxi.com/anthropic")
print(MINIMAX_API_KEY)

def load_prompt_file(filepath: str) -> str:
    """从md文件加载prompt内容"""
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            return f.read()
    except FileNotFoundError:
        return f"错误: Prompt文件未找到: {filepath}"
    except Exception as e:
        return f"错误: 读取Prompt文件失败: {e}"


def get_report_date_range() -> tuple:
    """根据配置获取日报统计日期范围
    返回 (since_date, until_date, description)
    - auto模式: 早上获取前一天,下午/晚上获取当天
    - manual模式: 获取最近DAILY_REPORT_DAYS天
    """
    now = datetime.now()
    current_hour = now.hour

    if DAILY_REPORT_MODE == "manual":
        since = (now - timedelta(days=DAILY_REPORT_DAYS)).strftime("%Y-%m-%d")
        until = (now + timedelta(days=1)).strftime("%Y-%m-%d")
        desc = f"最近{DAILY_REPORT_DAYS}天"
        return since, until, desc
    else:
        # auto模式
        if current_hour >= MORNING_HOUR and current_hour < EVENING_HOUR:
            # 早上，获取前一天数据
            since = (now - timedelta(days=1)).strftime("%Y-%m-%d")
            until = now.strftime("%Y-%m-%d")
            desc = "昨天"
        else:
            # 下午/晚上，获取当天数据
            since = now.strftime("%Y-%m-%d")
            until = (now + timedelta(days=1)).strftime("%Y-%m-%d")
            desc = "今天"
        return since, until, desc


def get_commits_in_range(repo_path: str, since: str, until: str, desc: str) -> str:
    """获取指定 Git 仓库的指定日期范围提交记录（全部作者）"""
    cmd = [
        "git", "-C", repo_path, "log",
        f"--since={since}", f"--until={until}",
        "--pretty=format:%h - %s (%an, %ad %H:%M)",
        "--date=short"
    ]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, encoding='utf-8', errors='replace')
        print(f"git repo_path:{repo_path},git result:{result}")
        if result.returncode == 0:
            return result.stdout.strip() or f"{desc}无提交记录"
        return f"获取Git日志失败: {result.stderr}"
    except FileNotFoundError:
        return "错误: 请确认已安装Git并加入PATH"
    except Exception as e:
        return f"执行Git命令异常: {e}"


def get_today_commits(repo_path: str) -> str:
    """获取指定 Git 仓库的当天提交记录（全部作者）"""
    since, until, desc = get_report_date_range()
    return get_commits_in_range(repo_path, since, until, desc)


def get_commits_with_hash_for_author_in_range(repo_path: str, author_name: str, author_email: str, since: str, until: str, desc: str) -> str:
    """获取指定仓库指定日期范围本人的提交记录（包含完整hash用于追溯）"""
    cmd = [
        "git", "-C", repo_path, "log",
        f"--since={since}", f"--until={until}",
        f"--author={author_name}",
        "--pretty=format:%H | %s | %an | %ad %H:%M",
        "--date=short"
    ]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, encoding='utf-8', errors='replace')
        if result.returncode == 0 and result.stdout.strip():
            lines = result.stdout.strip().split('\n')
            formatted = []
            for line in lines:
                parts = line.split(" | ", 3)
                if len(parts) == 4:
                    commit_hash = parts[0][:12]
                    message = parts[1]
                    author = parts[2]
                    date = parts[3]
                    formatted.append(f"- [{commit_hash}] {message} ({author}, {date})")
                else:
                    formatted.append(line)
            return "\n".join(formatted) if formatted else f"{desc}无提交记录"
        # 如果name没结果，尝试email
        cmd[-3] = f"--author={author_email}"
        result = subprocess.run(cmd, capture_output=True, text=True, encoding='utf-8', errors='replace')
        if result.returncode == 0 and result.stdout.strip():
            lines = result.stdout.strip().split('\n')
            formatted = []
            for line in lines:
                parts = line.split(" | ", 3)
                if len(parts) == 4:
                    commit_hash = parts[0][:12]
                    message = parts[1]
                    author = parts[2]
                    date = parts[3]
                    formatted.append(f"- [{commit_hash}] {message} ({author}, {date})")
                else:
                    formatted.append(line)
            return "\n".join(formatted) if formatted else f"{desc}无提交记录"
        return f"{desc}无提交记录"
    except Exception as e:
        return f"获取Git提交记录异常: {e}"


def get_today_commits_with_hash_for_author(repo_path: str, author_name: str, author_email: str) -> str:
    """获取指定仓库当日本人的提交记录（包含完整hash用于追溯）"""
    since, until, desc = get_report_date_range()
    return get_commits_with_hash_for_author_in_range(repo_path, author_name, author_email, since, until, desc)


def get_authors_in_range(repo_path: str, since: str, until: str) -> list:
    """获取指定日期范围Git提交中的所有唯一作者信息"""
    cmd = [
        "git", "-C", repo_path, "log",
        f"--since={since}", f"--until={until}",
        "--pretty=format:%an|%ae|%an",
        "--date=short"
    ]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, encoding='utf-8', errors='replace')
        if result.returncode == 0:
            lines = result.stdout.strip().split('\n') if result.stdout.strip() else []
            authors = {}
            for line in lines:
                parts = line.split("|")
                if len(parts) >= 3:
                    name, email, username = parts[0], parts[1], parts[2]
                    key = f"{name}|{email}"
                    if key not in authors:
                        authors[key] = {"name": name, "email": email, "git_name": username}
            return list(authors.values())
        return []
    except Exception as e:
        print(f"获取Git作者异常: {e}")
        return []


def get_today_authors(repo_path: str) -> list:
    """获取今日Git提交中的所有唯一作者信息"""
    since, until, desc = get_report_date_range()
    return get_authors_in_range(repo_path, since, until)


def get_all_git_authors() -> list:
    """获取所有项目在统计范围内的所有唯一Git作者"""
    since, until, desc = get_report_date_range()
    repo_map = {PLATFORM_DIR: PROJECT1_REPO_PATH, AGENT_DIR: PROJECT2_REPO_PATH}

    all_authors = {}
    for project, repo_path in repo_map.items():
        authors = get_authors_in_range(repo_path, since, until)
        for author in authors:
            key = f"{author['name']}|{author['email']}"
            if key not in all_authors:
                all_authors[key] = {
                    "name": author["name"],
                    "email": author["email"],
                    "projects": [project]
                }
            else:
                all_authors[key]["projects"].append(project)

    result = list(all_authors.values())
    return result


def get_diff_in_range(repo_path: str, since: str, until: str, author: str = None, max_lines: int = MAX_DIFF_LINES) -> str:
    """获取指定日期范围的代码变更（patch diff）"""
    cmd = [
        "git", "-C", repo_path, "log",
        f"--since={since}", f"--until={until}",
        "--patch",
        "--pretty=format:",
    ]

    if author:
        cmd.insert(2, f"--author={author}")

    try:
        result = subprocess.run(cmd, capture_output=True, text=True, encoding='utf-8', errors='replace')
        if result.returncode == 0:
            diff = result.stdout.strip()
            if not diff:
                return "该时间段无代码变更"
            lines = diff.split('\n')
            if len(lines) > max_lines:
                return "\n".join(lines[:max_lines]) + "\n... (diff 内容过长，已截断，仅显示前 {} 行)".format(max_lines)
            return diff
        return f"获取Git diff失败: {result.stderr}"
    except Exception as e:
        return f"执行Git diff命令异常: {e}"


def get_today_diff(repo_path: str, author: str = None, max_lines: int = MAX_DIFF_LINES) -> str:
    """获取当天的代码变更（patch diff）"""
    since, until, desc = get_report_date_range()
    return get_diff_in_range(repo_path, since, until, author, max_lines)


def read_local_plan_files(project: str = None) -> str:
    """从本地读取 Excel 或 Markdown 规划文件，合并为文本"""
    base_dir = PLAN_DATA_DIR
    if project:
        base_dir = os.path.join(PLAN_DATA_DIR, project)

    if base_dir and os.path.isdir(base_dir):
        files = []
        for f in os.listdir(base_dir):
            if f.endswith('.xlsx') or f.endswith('.md'):
                files.append(os.path.join(base_dir, f))
        if not files:
            return f"（未在 {project or '根目录'} 下找到任何 .xlsx 或 .md 文件）"
    else:
        return f"（未配置规划文件路径，请设置 PLAN_DATA_DIR）"

    content_parts = []
    for filepath in files:
        filename = os.path.basename(filepath)
        if filepath.endswith('.xlsx'):
            try:
                wb = openpyxl.load_workbook(filepath, data_only=True)
                sheet = wb.active
                lines = []
                for row in sheet.iter_rows(values_only=True):
                    row_str = "\t".join([str(c) if c is not None else "" for c in row])
                    lines.append(row_str)
                content_parts.append(f"### 文件: {filename} (Excel)\n" + "\n".join(lines))
            except Exception as e:
                content_parts.append(f"### 文件: {filename} (Excel) 读取失败: {e}")
        elif filepath.endswith('.md'):
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    md_content = f.read()
                content_parts.append(f"### 文件: {filename} (Markdown)\n{md_content}")
            except Exception as e:
                content_parts.append(f"### 文件: {filename} (Markdown) 读取失败: {e}")

    return "\n\n".join(content_parts)


def parse_excel_tasks(filepath: str) -> list:
    """解析Excel任务文件，返回结构化任务数据列表"""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(c).strip() if c else "" for c in rows[0]]
        col_map = {}
        for idx, h in enumerate(headers):
            h_lower = h.lower()
            if "模块" in h or "名称" in h:
                col_map["module"] = idx
            elif "功能" in h and "描述" in h:
                col_map["description"] = idx
            elif "描述" in h and "功能" not in h:
                col_map["description"] = idx
            elif "后端状态" in h:
                col_map["backend_status"] = idx
            elif "前端状态" in h:
                col_map["frontend_status"] = idx
            elif "后端负责人" in h:
                col_map["backend_owner"] = idx
            elif "前端负责人" in h:
                col_map["frontend_owner"] = idx
            elif "备注" in h:
                col_map["notes"] = idx
            elif "状态" in h and "backend" not in col_map and "frontend" not in col_map:
                if "backend_status" not in col_map:
                    col_map["backend_status"] = idx
                elif "frontend_status" not in col_map:
                    col_map["frontend_status"] = idx

        tasks = []
        for row_idx, row in enumerate(rows[1:], start=2):
            row_data = {}
            row_data["row"] = row_idx
            row_data["module"] = str(row[col_map.get("module", 0)]).strip() if col_map.get("module") is not None and col_map.get("module", 0) < len(row) else ""
            row_data["description"] = str(row[col_map.get("description", 1)]).strip() if col_map.get("description") is not None and col_map.get("description", 1) < len(row) else ""
            row_data["backend_status"] = str(row[col_map.get("backend_status", 6)]).strip() if col_map.get("backend_status") is not None and col_map.get("backend_status", 6) < len(row) else ""
            row_data["frontend_status"] = str(row[col_map.get("frontend_status", 9)]).strip() if col_map.get("frontend_status") is not None and col_map.get("frontend_status", 9) < len(row) else ""
            row_data["backend_owner"] = str(row[col_map.get("backend_owner", 5)]).strip() if col_map.get("backend_owner") is not None and col_map.get("backend_owner", 5) < len(row) else ""
            row_data["frontend_owner"] = str(row[col_map.get("frontend_owner", 8)]).strip() if col_map.get("frontend_owner") is not None and col_map.get("frontend_owner", 8) < len(row) else ""
            row_data["notes"] = str(row[col_map.get("notes", 11)]).strip() if col_map.get("notes") is not None and col_map.get("notes", 11) < len(row) else ""
            if row_data["module"] or row_data["description"]:
                tasks.append(row_data)
        return tasks
    except Exception as e:
        return []


def filter_person_tasks(tasks: list, person_name: str, role: str = "backend") -> list:
    """根据人名和角色过滤出属于该人员的待办任务"""
    filtered = []
    for task in tasks:
        owner = task.get("backend_owner") if role == "backend" else task.get("frontend_owner")
        # owner may have @ prefix like "@刘小辉" or be multiple people "@刘小辉 @张三"
        owner_clean = owner.replace("@", "").replace(" ", "") if owner else ""
        if person_name in owner or person_name in owner_clean:
            status = task.get("backend_status") if role == "backend" else task.get("frontend_status")
            if status and COMPLETED_STATUS not in status and status not in ["1.暂不考虑", "暂不考虑"]:
                filtered.append(task)
    return filtered[:MAX_PENDING_TASKS]


def match_tasks_with_commits(tasks: list, commits_text: str) -> list:
    """将任务列表与Git提交记录交叉验证，更新任务状态"""
    import re
    if not commits_text or not tasks:
        return tasks

    commit_lines = [line for line in commits_text.split("\n") if line.strip()]
    for task in tasks:
        task["matched_commits"] = []
        task["real_status"] = task.get("backend_status", "")

        keywords = []
        if task.get("module"):
            keywords.extend([w for w in task["module"].split() if len(w) > 2])
        if task.get("description"):
            keywords.extend([w for w in task["description"].split() if len(w) > 2])

        for line in commit_lines:
            for kw in keywords[:10]:
                if kw in line.lower():
                    task["matched_commits"].append(line.strip())
                    task["real_status"] = "进行中(已有提交)"
                    break
            if len(task["matched_commits"]) >= 3:
                break
    return tasks


def generate_task_llm_guide(task: dict, context_commits: list = None) -> str:
    """为单个任务生成LLM交互指南（简洁版）"""
    module = task.get("module", "")
    desc = task.get("description", "")
    status = task.get("backend_status", "") or task.get("frontend_status", "")
    notes = task.get("notes", "")
    matched = task.get("matched_commits", [])

    guide = f"**任务**: {module}"
    if desc:
        guide += f" - {desc[:50]}{'...' if len(desc) > 50 else ''}"

    guide += f"\n**当前状态**: {status}"

    if notes and notes not in ["无", "无备注"]:
        guide += f"\n**备注**: {notes[:30]}{'...' if len(notes) > 30 else ''}"

    guide += "\n**解决思路**: "
    if "待联调" in status:
        guide += "该任务需要前后端联调，建议先确认接口协议和数据格式，使用LLM辅助生成联调测试用例和Mock数据。"
    elif "开发中" in status or "进行中" in status:
        guide += "该任务正在开发中，建议利用LLM辅助代码实现，定期让LLM审查代码逻辑和边界情况。"
    else:
        guide += f"该任务状态为{status}，建议先理解需求背景，使用LLM辅助拆解任务步骤和生成代码框架。"

    guide += "\n**LLM交互提示**:"
    guide += f"\n- 需求理解: \"请帮我分析{module}模块的核心功能和实现要点\""
    if matched:
        commits_summary = matched[0][:30] if matched else ""
        guide += f"\n- 代码实现: \"基于已有提交'{commits_summary}...'，请帮我继续实现剩余功能\""
    else:
        guide += f"\n- 代码实现: \"请为{module}模块生成代码实现，需求：{desc[:30]}{'...' if len(desc) > 30 else ''}\""
    guide += "\n- 自测验证: \"请为该功能生成单元测试用例，覆盖正常和异常场景\""

    return guide


def generate_pending_tasks_section(person_name: str, project_roles: dict) -> str:
    """生成待办任务及LLM交互指南section"""
    parts = []

    for project, roles_str in project_roles.items():
        project_dir = os.path.join(PLAN_DATA_DIR, project)
        if not os.path.isdir(project_dir):
            continue

        all_tasks = []
        for f in os.listdir(project_dir):
            if f.endswith(".xlsx"):
                filepath = os.path.join(project_dir, f)
                tasks = parse_excel_tasks(filepath)
                all_tasks.extend(tasks)

        for role in roles_str.split(","):
            role = role.strip()
            if role == "developer" or role == "架构师":
                tasks_backend = filter_person_tasks(all_tasks, person_name, "backend")
                tasks_frontend = filter_person_tasks(all_tasks, person_name, "frontend")
                person_tasks = tasks_backend if tasks_backend else tasks_frontend
            elif role == "team_leader":
                person_tasks = filter_person_tasks(all_tasks, person_name, "backend")
                person_tasks.extend(filter_person_tasks(all_tasks, person_name, "frontend"))
            else:
                continue

            if not person_tasks:
                continue

            repo_path = get_repo_path_for_project(project)
            from datetime import datetime
            now = datetime.now()
            current_hour = now.hour
            if current_hour >= 6 and current_hour < 12:
                since = (now - timedelta(days=1)).strftime("%Y-%m-%d")
            else:
                since = now.strftime("%Y-%m-%d")
            until = (now + timedelta(days=1)).strftime("%Y-%m-%d")
            commits = get_commits_with_hash_for_author_in_range(repo_path, person_name, person_name, since, until, "统计")

            person_tasks = match_tasks_with_commits(person_tasks, commits)

            project_name = get_project_name(project)
            parts.append(f"**{project_name}（{project}）**")
            for i, task in enumerate(person_tasks, 1):
                guide = generate_task_llm_guide(task)
                parts.append(f"{i}. {guide}")
            parts.append("")

    if not parts:
        return "（暂无待办任务）"

    return "\n".join(parts)


def get_repo_path_for_project(project: str) -> str:
    """获取项目对应的仓库路径"""
    if project == PLATFORM_DIR:
        return PROJECT1_REPO_PATH
    elif project == AGENT_DIR:
        return PROJECT2_REPO_PATH
    return None


def get_project_name(project: str) -> str:
    """获取项目显示名称"""
    if project == PLATFORM_DIR:
        return PROJECT1_NAME
    elif project == AGENT_DIR:
        return PROJECT2_NAME
    return project


def generate_personal_context(member: dict, all_members: list) -> str:
    """生成个人日报上下文"""
    plan_files = {}
    git_commits = {}
    personal_diffs = {}
    personal_commits_with_hash = {}
    project_role_map = member["project_roles"]

    for project in project_role_map.keys():
        plan_files[project] = read_local_plan_files(project)
        repo_path = get_repo_path_for_project(project)
        git_commits[project] = get_today_commits(repo_path)
        # 获取本人的提交记录（含hash用于追溯）
        personal_commits_with_hash[project] = get_today_commits_with_hash_for_author(repo_path, member["git_name"], member["email"])

        diff = get_today_diff(repo_path, author=member["git_name"])
        if diff.startswith("今日无代码变更") or diff.startswith("获取Git diff失败"):
            diff_email = get_today_diff(repo_path, author=member["email"])
            if not diff_email.startswith("获取Git diff失败"):
                diff = diff_email
        personal_diffs[project] = diff

    plan_parts = []
    for project in project_role_map.keys():
        plan_parts.append(f"**{get_project_name(project)}（{project}目录）**\n{plan_files[project]}")

    commit_parts = []
    diff_parts = []
    personal_commit_parts = []
    for project in project_role_map.keys():
        commit_parts.append(f"**{get_project_name(project)}**\n{git_commits[project]}")
        diff_parts.append(f"**{get_project_name(project)}**\n{personal_diffs[project]}")
        personal_commit_parts.append(f"**{get_project_name(project)}**\n{personal_commits_with_hash[project]}")

    # 收集所有角色用于prompt
    all_roles = set()
    for roles_str in project_role_map.values():
        for r in roles_str.split(","):
            all_roles.add(r.strip())

    today = datetime.now().strftime("%Y-%m-%d")

    today = datetime.now().strftime("%Y-%m-%d")

    # 生成待办任务section
    pending_section = generate_pending_tasks_section(member["name"], member["project_roles"])

    context = f"""日期：{today}

### 项目规划（本地文件）
{chr(10).join(plan_parts)}

### 今日 Git 提交记录（全部成员）

{chr(10).join(commit_parts)}

### 本人今日提交记录（含提交号，用于追溯）

{chr(10).join(personal_commit_parts)}

### 本人今日代码变更 (Diff)

{chr(10).join(diff_parts)}

### 今日待办任务及LLM交互指南

{pending_section}

人员信息：{member['name']}，Git用户名：{member['git_name']}，邮箱：{member['email']}"""
    return context, ",".join(all_roles)


def generate_project_context(project: str, all_members: list) -> str:
    """生成项目日报上下文"""
    repo_path = get_repo_path_for_project(project)
    all_commits = get_today_commits(repo_path)
    plan_text = read_local_plan_files(project)

    member_diffs = []
    for member in all_members:
        if project in member["project_roles"]:
            diff = get_today_diff(repo_path, author=member["git_name"])
            if diff.startswith("今日无代码变更") or diff.startswith("获取Git diff失败"):
                diff_email = get_today_diff(repo_path, author=member["email"])
                if not diff_email.startswith("获取Git diff失败"):
                    diff = diff_email
            roles_str = member["project_roles"][project]
            role_display = get_roles_display(roles_str)
            member_diffs.append(f"**{member['name']}（{role_display}）**\n{diff}")

    today = datetime.now().strftime("%Y-%m-%d")

    context = f"""日期：{today}

### 项目：{get_project_name(project)}

### 项目规划（{project}目录）
{plan_text}

### 今日 Git 提交记录（全部成员）
{all_commits}

### 各成员代码变更
{chr(10).join(member_diffs)}

项目角色说明：team_leader=团队负责人, developer=开发人员, architecture=架构师"""
    return context


def generate_report(context: str, prompt_file: str, **format_kwargs) -> str:
    """调用大模型生成报告"""
    client = anthropic.Anthropic(
        base_url=MINIMAX_BASE_URL,
        api_key=MINIMAX_API_KEY,
    )

    prompt_template = load_prompt_file(prompt_file)
    system_prompt = prompt_template.format(**format_kwargs)

    response = client.messages.create(
        model=MODEL_NAME,
        max_tokens=MAX_TOKENS,
        system=system_prompt,
        messages=[{"role": "user", "content": [{"type": "text", "text": context}]}],
        temperature=TEMPERATURE,
    )

    text_blocks = [block.text for block in response.content if block.type == "text"]
    return "\n".join(text_blocks)


def get_role_display(role: str) -> str:
    """获取单个角色中文显示"""
    role_map = {
        "team_leader": "团队负责人",
        "developer": "开发人员",
        "architecture": "架构师"
    }
    return role_map.get(role, role)


def get_roles_display(roles_str: str) -> str:
    """获取多个角色中文显示（逗号分隔）"""
    roles = [r.strip() for r in roles_str.split(",")]
    displays = [get_role_display(r) for r in roles]
    return "、".join(displays)


def discover_team_members_from_git() -> list:
    """从Git提交记录自动发现团队成员"""
    all_authors = {}
    repo_map = {PLATFORM_DIR: PROJECT1_REPO_PATH, AGENT_DIR: PROJECT2_REPO_PATH}

    for project, repo_path in repo_map.items():
        authors = get_today_authors(repo_path)
        for author in authors:
            key = f"{author['name']}|{author['email']}"
            if key not in all_authors:
                all_authors[key] = {
                    "name": author["name"],
                    "git_name": author["git_name"],
                    "email": author["email"],
                    "project_roles": {}
                }
            if project not in all_authors[key]["project_roles"]:
                all_authors[key]["project_roles"][project] = DEFAULT_ROLE

    return list(all_authors.values())


def get_team_members() -> list:
    """获取团队成员列表，根据配置决定来源"""
    if TEAM_MEMBER_SOURCE == "git":
        members = discover_team_members_from_git()
        print(f"[INFO] 从Git自动发现 {len(members)} 名成员")
        return members
    else:
        return TEAM_MEMBERS


if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding='utf-8')

    # 获取报表统计日期范围
    since, until, date_desc = get_report_date_range()
    print(f">>> 正在提取{date_desc}数据并调用 {MODEL_NAME} 生成日报...\n")

    today = datetime.now()
    year_month = today.strftime("%Y%m")
    day = today.strftime("%d")
    date_str = today.strftime("%Y-%m-%d")

    # 输出目录结构：output/YYYYMM/DD/by_person/daily/ 和 output/YYYYMM/DD/by_project/daily/
    output_base = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "output", year_month, day)
    person_output_dir = os.path.join(output_base, "by_person", "daily")
    project_output_dir = os.path.join(output_base, "by_project", "daily")
    os.makedirs(person_output_dir, exist_ok=True)
    os.makedirs(project_output_dir, exist_ok=True)

    # 获取团队成员（根据配置决定来源）
    team_members = get_team_members()

    # 收集待发送的报告信息（先生成报告，暂不发送邮件）
    personal_reports = []  # [(member, filepath), ...]
    project_reports = []   # [(project_name, filepath, recipients), ...]

    # 1. 生成个人日报
    print("=== 生成个人日报 ===")
    for member in team_members:
        ctx, roles_str = generate_personal_context(member, team_members)
        role_display = get_roles_display(roles_str)
        print(f"--- 为 {member['name']}（{role_display}）生成日报 ---")
        report = generate_report(
            ctx,
            PERSONAL_DAILY_PROMPT_FILE,
            MY_NAME=member['name'],
            MY_ROLE=role_display,
            MY_GIT_NAME=member['git_name'],
            MY_EMAIL=member['email']
        )
        filename = f"{member['name']}_日报_{date_str}_{date_desc}.md"
        filepath = os.path.join(person_output_dir, filename)
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(report)
        print(f"[OK] {member['name']} 日报已保存")
        personal_reports.append((member, filepath))

    # 2. 生成项目日报
    print("\n=== 生成项目日报 ===")
    for project in [PLATFORM_DIR, AGENT_DIR]:
        print(f"--- 为 {get_project_name(project)} 生成日报 ---")
        ctx = generate_project_context(project, team_members)
        report = generate_report(
            ctx,
            PROJECT_DAILY_PROMPT_FILE,
            PROJECT_NAME=get_project_name(project),
            PROJECT_DIR=project
        )
        filename = f"{get_project_name(project)}_日报_{date_str}_{date_desc}.md"
        filepath = os.path.join(project_output_dir, filename)
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(report)
        print(f"[OK] {get_project_name(project)} 日报已保存")
        recipients = PROJECT_REPORT_RECIPIENTS.get(project, DEFAULT_PROJECT_RECIPIENTS)
        if recipients:
            project_reports.append((get_project_name(project), filepath, recipients))

    # 3. 邮件发送确认（如果启用邮件）
    if EMAIL_ENABLED and (personal_reports or project_reports):
        # 重置邮件发送频率计数
        reset_email_rate_limit()

        # 获取所有Git作者列表
        git_authors = get_all_git_authors()

        # 收集所有收件人
        if TEAM_MEMBER_EMAIL_MODE == "select":
            # 从git作者中选择发送目标，显示所有git用户作为候选
            selected_emails = select_recipients_from_git_authors(git_authors)
            if not selected_emails:
                print("[INFO] 未选择任何收件人，跳过发送个人报告")
                personal_reports_filtered = []
                personal_recipients = []
            else:
                personal_recipients = selected_emails
                # 过滤 personal_reports 只保留选中邮箱的成员
                personal_reports_filtered = [(member, filepath) for member, filepath in personal_reports if member['email'] in selected_emails]
        else:
            # 发送给所有成员
            personal_recipients = [member['email'] for member, _ in personal_reports if TEAM_MEMBER_SEND_EMAIL]
            personal_reports_filtered = personal_reports

        # 为每个项目选择项目报告收件人
        project_report_recipients = {}  # {project_name: [emails]}
        for project in [PLATFORM_DIR, AGENT_DIR]:
            project_name = get_project_name(project)
            # 获取该项目相关的Git作者（按项目过滤）
            project_authors = [a for a in git_authors if project in a.get("projects", [])]
            if project_authors:
                # 转换为 member 格式
                project_members = [{"name": a["name"], "email": a["email"]} for a in project_authors]
                recipients = select_project_report_recipients(project_name, project_members)
                if recipients:
                    project_report_recipients[project] = recipients

        # 确认发件人和收件人
        confirmed, sender_email, all_recipients = confirm_email_sender_and_recipients(
            git_authors,
            EMAIL_SMTP_USER,
            personal_recipients=personal_recipients if personal_recipients else None,
            project_recipients=list(set(email for emails in project_report_recipients.values() for email in emails)) if project_report_recipients else None
        )

        if confirmed:
            # 发送个人报告邮件
            print(f"[DEBUG] personal_reports_filtered: {[(m['name'], m['email']) for m, _ in personal_reports_filtered]}")
            for member, filepath in personal_reports_filtered:
                if TEAM_MEMBER_SEND_EMAIL:
                    send_personal_report_email(member, filepath, "日报")
            # 发送项目报告邮件
            for project_name, filepath, _ in project_reports:
                project_key = PLATFORM_DIR if "数据平台" in project_name else AGENT_DIR
                recipients = project_report_recipients.get(project_key, [])
                if recipients:
                    print(f"[DEBUG] 项目报告收件人: project={project_name}, recipients={recipients}")
                    send_project_report_email(project_name, filepath, "日报", recipients)
                else:
                    print(f"[DEBUG] 项目报告 {project_name} 未选择收件人，跳过发送")
        else:
            print("[INFO] 邮件发送已取消")

    print(f"\n[INFO] 统计范围: {date_desc} ({since} ~ {until})")
    print(f"[INFO] 输出目录: {output_base}")
    print(f"[INFO] 个人日报: {person_output_dir}")
    print(f"[INFO] 项目日报: {project_output_dir}")